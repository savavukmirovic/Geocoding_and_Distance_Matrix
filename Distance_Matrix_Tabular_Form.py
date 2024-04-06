import openpyxl
import os
from dotenv import load_dotenv

load_dotenv()

wdir = os.environ.get("USER_PATH")

wb = openpyxl.load_workbook(wdir, data_only=True)
ws = wb["Distance_Matrix"]
ws_tabular_form = wb["Distance_Matrix_Tabular_Form"]

address_excel = []
place_excel = []
zip_excel = []
for row in range(ws.max_row-1):
    address_excel.append(ws.cell(2 + row, 1).value)
    place_excel.append(ws.cell(2 + row, 2).value)
    zip_excel.append(ws.cell(2 + row, 3).value)

counter = 0
starting_row_address_excel = 2
starting_column_address_excel = 1
starting_row_place_excel = 2
starting_column_place_excel = 2
starting_row_zip_excel = 2
starting_column_zip_excel = 3

for from_address in range(len(address_excel)):
    for to_address in range(len(address_excel)):
        ws_tabular_form.cell(starting_row_address_excel + counter,
                             starting_column_address_excel, address_excel[from_address])
        ws_tabular_form.cell(starting_row_place_excel + counter, starting_column_place_excel, place_excel[from_address])
        ws_tabular_form.cell(starting_row_zip_excel + counter, starting_column_zip_excel, zip_excel[from_address])

        ws_tabular_form.cell(starting_row_address_excel + counter, 4, address_excel[to_address])
        ws_tabular_form.cell(starting_row_place_excel + counter, 5, place_excel[to_address])
        ws_tabular_form.cell(starting_row_place_excel + counter, 6, zip_excel[to_address])
        ws_tabular_form.cell(starting_row_address_excel + counter, 7, ws.cell(from_address + 2, to_address + 7).value)
        counter += 1
        
ws_tabular_form.cell(1, 1, "Address")
ws_tabular_form.cell(1, 2, "Place")
ws_tabular_form.cell(1, 3, "ZIP")
ws_tabular_form.cell(1, 4, "Address")
ws_tabular_form.cell(1, 5, "Place")
ws_tabular_form.cell(1, 6, "ZIP")
ws_tabular_form.cell(1, 7, "Distance_km")

wb.save(wdir)
wb.close()
