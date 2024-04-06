import json
import requests
import openpyxl
import os
from dotenv import load_dotenv

load_dotenv()

wdir = os.environ.get("USER_PATH")

wb = openpyxl.load_workbook(wdir, data_only=True)
ws = wb["Distance_Matrix"]
max_row = ws.max_row

y_coordinate = []
for y_row in range(max_row - 1):
    cellitem1 = ws.cell(2 + y_row, 4).value
    y_coordinate.append(cellitem1)

x_coordinate = []
for x_row in range(max_row - 1):
    cellitem2 = ws.cell(2 + x_row, 5).value
    x_coordinate.append(cellitem2)

for row in range(2, ws.max_row + 1):
    ws.cell(1, 7 + row - 2, ws.cell(row, 1).value)

counter = 0
distance_to_excel = []
for first_obj_coord in range(len(y_coordinate)):
    from_first_obj_coord = []
    for second_obj_coord in range(0, len(y_coordinate)):
        x1 = float(y_coordinate[first_obj_coord])
        y1 = float(x_coordinate[first_obj_coord])
        x2 = float(y_coordinate[second_obj_coord])
        y2 = float(x_coordinate[second_obj_coord])

        r = requests.get(f"http://router.project-osrm.org/route/v1/car/{y1},{x1};{y2},{x2}?overview=false""")
        routes = json.loads(r.content)
        route = routes.get("routes")[0]
        rastojanjekm = round(route["legs"][0]["distance"]/1000, 2)
        print(f'Distance: {x1, y1}-{x2, y2}: {rastojanjekm}')
        from_first_obj_coord.append(rastojanjekm)
    distance_to_excel.append(from_first_obj_coord)

    start_from_row = 2 + counter
    start_from_column = 7
    for i in range(1):
        for j in range(0, max_row - 1):
            ws.cell(start_from_row, start_from_column + j, distance_to_excel[counter][j])
    counter += 1
    wb.save(wdir)
wb.close()
