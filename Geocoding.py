import openpyxl
from geopy.geocoders import Nominatim
from dotenv import load_dotenv
import os


def find_coordinate(address):
    """This function finds the latitude and longitude of a given place"""
    try:
        location = geolocator.geocode(address, exactly_one=True)
        if not location:
            result = ("ADDRESS NOT FOUND", "ADDRESS NOT FOUND")
        else:
            result = (location.longitude, location.latitude)
    except FileNotFoundError:
        result = ("ADDRESS NOT FOUND", "ADDRESS NOT FOUND")
    return result


load_dotenv()

geolocator = Nominatim(user_agent=os.environ.get("USER_AGENT"))
wdir = os.environ.get("USER_PATH")


wb = openpyxl.load_workbook(wdir, data_only=True)
ws = wb["Geocoding"]
ws.cell(1, 4, "GPS-Latitude")
ws.cell(1, 5, "GPS-Longitude")

addressexcel = []
for row in range(2, ws.max_row + 1):
    addressexcel.append(ws.cell(row, 1).value)


place_coordinates = []
for addr in addressexcel:
    place_coordinates.append(find_coordinate(address=addr))

    place_coordinates_x = []
    place_coordinates_y = []
    for coordinate in place_coordinates:
        place_coordinates_x.append(coordinate[0])
        place_coordinates_y.append(coordinate[1])

    starting_column_exceladdress_y = 4
    starting_column_exceladdress_x = 5
    for row in range(len(place_coordinates_x)):
        ws.cell(row + 2, starting_column_exceladdress_x, place_coordinates_x[row])
        ws.cell(row + 2, starting_column_exceladdress_y, place_coordinates_y[row])

    wb.save(wdir)
wb.close()
