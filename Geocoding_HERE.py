import openpyxl
import requests
import os
from dotenv import load_dotenv


def coordinate(address):
    """This function finds the latitude and longitude of a given place"""
    url = "https://discover.search.hereapi.com/v1/geocode"

    parameters = {
        "q": f"{address}",
        "apiKey": F"{os.environ.get('HERE_API_KEY')}"
    }
    try:
        response = requests.get(url=url, params=parameters)
        response = response.json()
        print(response)
        result = (response['items'][0]['position']['lng'], response['items'][0]['position']['lat'])
    except IndexError or KeyError:
        result = ("ADDRESS NOT FOUND", "ADDRESS NOT FOUND")
    return result


load_dotenv()

wdir = os.environ.get("USER_PATH")

wb = openpyxl.load_workbook(wdir, data_only=True)
ws = wb["Geocoding"]
ws.cell(1, 4, "GPS-Latitude")
ws.cell(1, 5, "GPS-Longitude")

addressexcel = []
for row in range(2, ws.max_row + 1):
    addressexcel.append(ws.cell(row, 1).value)


place_coordinates = []
for addr in addressexcel:
    place_coordinates.append(coordinate(address=addr))

    place_coordinates_x = []
    place_coordinates_y = []
    for coordinate in place_coordinates:
        place_coordinates_x.append(coordinate[0])
        place_coordinates_y.append(coordinate[1])

    starting_column_exceladdress_y = 4
    starting_column_exceladdress_x = 5
    for row in range(len(place_coordinates_x)):
        ws.cell(row + 2, starting_column_exceladdress_x, place_coordinates_x[row])
        ws.cell(row + 2, starting_column_exceladdress_y, place_coordinates_y[row])

    wb.save(wdir)
wb.close()
